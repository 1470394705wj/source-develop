# encoding='utf-8'
"""
ModuleName: diff比对csv文件内容的共通方法
Description:
date: 2022/9/5 9:24
@author caoting
@version 1.0
@since Python 3.6
"""
import json
from copy import copy

from gevent import monkey
monkey.patch_all()
# from common.constants import result_log_dir
from openpyxl.styles import PatternFill
import datacompy
import openpyxl
import os
import hashlib
from config.data_test.mds_respond.cfgs import SERVERS_Cfg,SelfSubtype_code_Cfg
from common.public_api import get_listcode_snap
try:
    from base_data.data_test.mds_respond import setting
except:
    pass
from common.yaml_utils import YamlUtils
from common.constants import BASE_DATA_DIR, DATA_DIR, REPORT_DIR

import itertools
import datetime
from concurrent.futures import ProcessPoolExecutor
import socket,requests
import pandas as pd
from common.data_test_base import DataTestBase


class Compare_csv_diff(DataTestBase):
    def sh1_compare_comm_fun(self):
        #第二迭代多进程比对
        #1.读取outputs/sh1/sh1_yyyy-mm-dd.txt 落地文件的路径
        #list_outcsv_path=[(path1,path2),(path1,path3)]
        list_outcsv_path=Compare_csv_diff().get_sh1_diffconf_dir()
        for path in list_outcsv_path:
            diff_path1=path[0]
            diff_path2=path[1]
            nowdate = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            api_types = os.listdir(diff_path1)  # 获取所有的接口类型
            # 获得市场名# r'E:\mds_OOP_autotest\outdata\sh1\2022年08月22日14时10分28秒\previewyunhq.sse.com.cn_32041'
            market = str(diff_path1.split('\\')[-3])
            t_list=[]
            for type in api_types:
                tupl1=(type,market,nowdate,diff_path1,diff_path2)
                t_list.append(tupl1)

            #多进程
            with ProcessPoolExecutor(max_workers=5) as pool:
                res=pool.map(Compare_csv_diff().tuple_param, t_list)



    def tuple_param(self,t_list):
        return Compare_csv_diff().diff_fun_apitype(t_list[0],t_list[1],t_list[2],t_list[3],t_list[4])


    #第二迭代多进程比对
    def diff_fun_apitype(self, apitype, market, nowdate, diff_path1, diff_path2):
        if apitype != 'monitor':  # monitor接口比较特殊,不做对比
            list_csvfile1 = Compare_csv_diff().return_api_csv(diff_path1, apitype)  # 获取第一个ip文件夹目录下的所有csv文件
            #print("apitype=",apitype,list_csvfile1)
            list_csvfile2 = Compare_csv_diff().return_api_csv(diff_path2, apitype)
            if len(list_csvfile1) > 0 and len(list_csvfile2) > 0:  # 确保接口类型下有csv文件，文件数大于0
                # print("list_csvfile1=",list_csvfile1)
                os.makedirs(f'{REPORT_DIR}/data_test/mds_respond/{market}', exist_ok=True)
                out_path = f'{REPORT_DIR}/data_test/mds_respond/{market}' + "\\" + nowdate
                if not os.path.exists(out_path):
                    os.makedirs(out_path)
                if not os.path.exists(out_path + "\\" + apitype):
                    os.makedirs(out_path + "\\" + apitype)
                rept_path = out_path + "\\" + apitype + "\\"
                # print("rept_path=" + rept_path)
                # 创建sh1市场输出报告路径report\sh1\date\apitype\
                same_count = 0  # 统计该接口 csv比对文件一致的场景数
                different_count = 0  # 统计该接口 csv比对文件不一致的场景数
                for fileapi_a1 in list_csvfile1:  # 判断两个ip目录下名字相同的csv文件做比对
                    for fileapi_b1 in list_csvfile2:
                        if fileapi_a1 == fileapi_b1:
                            apifilenm1 = os.path.join(os.path.join(diff_path1, apitype), fileapi_a1)
                            apifilenm2 = os.path.join(os.path.join(diff_path2, apitype), fileapi_b1)
                            m1 = Compare_csv_diff().file_md5(apifilenm1)
                            m2 = Compare_csv_diff().file_md5(apifilenm2)
                            if m1 == m2:  # 文件md5转码值相等，不用对比操作内容
                                same_count = same_count + 1
                            else:
                                #两个csv文件比对不一致写入execl
                                #计算比对不同的用例数
                                different_count = different_count + 1
                                # csv文件内容不一致就比对内容输出报告
                                # 读setting.py配置diff_html，diff_csv
                                # 测试报告输出到report\
                                diff_fnm = str(apifilenm1.split('\\')[-1])
                                diff_path_fm = rept_path + str(diff_fnm.split('.')[0] + 'diff.xlsx')
                                # 创建diff_execl文件作为输出测试报告文件
                                wb = openpyxl.Workbook()
                                # ws = wb.active
                                # sheet = wb.create_sheet('sheet', 0)  # 创建一个sheet，用于存放数据
                                wb.save(diff_path_fm)  # 保存到excel文件
                                version = str(diff_fnm.split('_')[0])  # 获取version
                                key = Compare_csv_diff().get_apiversion_key(apitype, version)
                                #多进程
                                # tuple1=(apifilenm1, apifilenm2, diff_path_fm, key)
                                # list_tup.append(tuple1)
                                # with ProcessPoolExecutor(max_workers=6) as pool:
                                #     res = pool.map(Compare_csv_diff().diff_compare_tup, list_tup)
                                Compare_csv_diff().data_compy(apifilenm1, apifilenm2, diff_path_fm, key)

        # res_str = f'{apitype}接口csv文件比对一致的用例数有{same_count}，不一致的用例数有{different_count}'
        # with open(result_log_dir, 'a') as file_handle:  # .txt可以不自己新建,代码会自动新建
        #     file_handle.write(res_str)  # 写入
        #     file_handle.write('\n')

        log_json = copy(self.logTemplate)
        log_json['_dataTime'] = datetime.datetime.now().__str__()
        log_json['_metric'] = f'{apitype}接口csv文件比对一致的用例数有{same_count}，不一致的用例数有{different_count}'
        self.logUtils.logger.info(json.dumps(log_json, ensure_ascii=False))



    def diff_compare_tup(self,list_tup):
        return Compare_csv_diff().data_compy(list_tup[0], list_tup[1], list_tup[2], list_tup[3])



    #读取落地配置文件中的各个市场的落地路径
    def get_sh1_diffconf_dir(self):
        list_dirpath = []
        #sh1市场落地文件路径
        to_date = str(datetime.date.today())
        sh1_fnm='sh1_'+to_date+'.txt'
        sh1_csv_dir=os.path.join(DATA_DIR, f'data_test/mds_respond/sh1/{sh1_fnm}')
        with open(sh1_csv_dir, 'r', encoding='utf-8') as f:
            for line in f:
                list_dirpath.append(line.replace('\n', ''))
        list_twocsvpath = list(itertools.combinations(list_dirpath, 2))
        return list_twocsvpath



    #第一迭代单进程 执行比对测试用例
    def act_diff_case(self):
        #读取配置文件中的落地文件路径
        list_twocsvpath=Compare_csv_diff().get_sh1_diffconf_dir()
        for i in range(len(list_twocsvpath)):
            #接口数据比对
            out_path=Compare_csv_diff().diff_csv_fun(list_twocsvpath[i][0], list_twocsvpath[i][1])
            #比对trd1，snap,line 横向比较成交量总和
            #Compare_csv_diff().snap_trd1_line_sumvolume(out_path)
            #trd2
            #Compare_csv_diff().snap_trd2_line_sumvolume(out_path)


    #snap_trd1_line sumvolume 横向成交量总和对比
    def snap_trd1_line_sumvolume(self,out_path):
        #根据yml文件读取code，写入列表liscode中
        listcode= get_listcode_snap(SelfSubtype_code_Cfg)
        diff_volum_fm = out_path + r'\trd1\diff_volume_trd1.xlsx'
        listxlsx=[]
        reportv1_title = ["host", "version", "code", "snap_volume", "line_volume", "trd1_volume"]
        listxlsx.append(reportv1_title)
        listversion = ['v1', 'v2']
        for version in listversion:
            for code in listcode:
                for server in SERVERS_Cfg:
                    if server["use"] == 1:
                    # 分离ip+端口，创建目录
                        host = server["host"]
                        # 接口请求获取volume
                        line_url = host + '/' + version + '/sh1/line/' + code + '?select=volume&begin=0&end=-1'
                        snap_url = host + '/' + version + '/sh1/snap/' + code + '?select=volume'
                        trd1_url = host + '/' + version + '/sh1/trd1/' + code + '?select=volume&begin=0&end=-1'
                        # 接口请求
                        res_get_line = requests.get(line_url)
                        res_get_snap = requests.get(snap_url)
                        res_get_trd1 = requests.get(trd1_url)
                        if res_get_line.status_code == 200 and res_get_snap.status_code == 200 and res_get_trd1.status_code == 200:
                            sum_line = 0
                            sum_trd1 = 0
                            line_list = res_get_line.json()["line"]
                            #print("line_list=",line_list)
                            for line in line_list:
                                sum_line += line[0]
                            trd1_list = res_get_trd1.json()["trd1"]
                            #print("trd1_list=",trd1_list)
                            for trd1 in trd1_list:
                                sum_trd1 += trd1[0]
                            snap_volume = res_get_snap.json()["snap"][0]
                            listvolume=[host,version,code,snap_volume,sum_line,sum_trd1]
                            listxlsx.append(listvolume)
        Compare_csv_diff().writetoxlsx(diff_volum_fm, listxlsx)
        Compare_csv_diff().color_volume(diff_volum_fm)



    def writetoxlsx(self,diff_volum_fm,listxlsx):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成交量总和比较'
        for r in range(len(listxlsx)):
            for c in range(len(listxlsx[0])):
                ws.cell(r + 1, c + 1).value = listxlsx[r][c]
                # excel中的行和列是从1开始计数的，所以需要+1
        wb.save(diff_volum_fm)  # 注意，写入后一定要保存
        print("成功写入文件: " + diff_volum_fm + " !")
        return  1


    # snap_trd1_line sumvolume 横向成交量总和对比
    def snap_trd2_line_sumvolume(self, out_path):
            # 根据yml文件读取code，写入列表liscode中
            listcode = get_listcode_snap(SelfSubtype_code_Cfg)
            diff_volum_fm = out_path + r'\trd2\diff_volume_trd2.xlsx'
            listxlsx = []
            reportv1_title = ["host", "version", "code", "snap_volume", "line_volume", "trd1_volume"]
            listxlsx.append(reportv1_title)
            listversion = ['v1', 'v2']
            for version in listversion:
                for code in listcode:
                    for server in SERVERS_Cfg:
                        if server["use"] == 1:
                            # 分离ip+端口，创建目录
                            host = server["host"]
                            # 接口请求获取volume
                            line_url = host + '/' + version + '/sh1/line/' + code + '?select=volume&begin=0&end=-1'
                            snap_url = host + '/' + version + '/sh1/snap/' + code + '?select=volume'
                            trd2_url = host + '/' + version + '/sh1/trd2/' + code + '?select=volume&begin=0&end=-1'
                            # print("trd1_url=", trd1_url,snap_url)
                            # 接口请求
                            res_get_line = requests.get(line_url)
                            res_get_snap = requests.get(snap_url)
                            res_get_trd2 = requests.get(trd2_url)
                            if res_get_line.status_code == 200 and res_get_snap.status_code == 200 and res_get_trd2.status_code == 200:
                                sum_line = 0
                                sum_trd2 = 0
                                line_list = res_get_line.json()["line"]
                                # print("line_list=",line_list)
                                for line in line_list:
                                    sum_line += line[0]
                                trd2_list = res_get_trd2.json()["trd2"]
                                for trd2 in trd2_list:
                                    sum_trd2 += trd2[1]
                                snap_volume = res_get_snap.json()["snap"][0]
                                # print(f'snap_volume={snap_volume},sum_line={sum_line},sum_trd1={sum_trd1}')
                                listvolume = [host, version, code, snap_volume, sum_line, sum_trd2]
                                # print("listvolume=",listvolume)
                                listxlsx.append(listvolume)
            Compare_csv_diff().writetoxlsx(diff_volum_fm, listxlsx)
            Compare_csv_diff().color_volume(diff_volum_fm)


    #比对csv文件共通方法
    def diff_csv_fun(self,path1,path2):
        api_types = os.listdir(path1)  # 获取所有的接口类型
        nowdate = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        # 获得市场名# r'E:\mds_OOP_autotest\outdata\sh1\2022年08月22日14时10分28秒\previewyunhq.sse.com.cn_32041'
        market = str(path1.split('\\')[-3])
        # print("market=",market)
        out_path=''
        for apitype in api_types:
            if apitype != 'monitor':  # monitor接口比较特殊,不做对比
                list_csvfile1 = Compare_csv_diff().return_api_csv(path1, apitype)  # 获取第一个ip文件夹目录下的所有csv文件
                list_csvfile2 = Compare_csv_diff().return_api_csv(path2, apitype)
                if len(list_csvfile1) > 0 and len(list_csvfile2) > 0:  # 确保接口类型下有csv文件，文件数大于0
                    # print("list_csvfile1=",list_csvfile1)
                    os.makedirs(f'{REPORT_DIR}/data_test/mds_respond/{market}', exist_ok=True)
                    out_path = f'{REPORT_DIR}/data_test/mds_respond/{market}' + "\\" + nowdate
                    if not os.path.exists(out_path):
                        os.makedirs(out_path)
                    if not os.path.exists(out_path + "\\" + apitype):
                        os.makedirs(out_path + "\\" + apitype)
                    rept_path = out_path + "\\" + apitype + "\\"
                    # print("rept_path=" + rept_path)
                    # 创建sh1市场输出报告路径report\sh1\date\apitype\
                    same_count = 0  # 统计该接口 csv比对文件一致的场景数
                    different_count = 0  # 统计该接口 csv比对文件不一致的场景数
                    for fileapi_a1 in list_csvfile1:  # 判断两个ip目录下名字相同的csv文件做比对
                        for fileapi_b1 in list_csvfile2:
                            if fileapi_a1 == fileapi_b1:
                                apifilenm1 = os.path.join(os.path.join(path1, apitype), fileapi_a1)
                                apifilenm2 = os.path.join(os.path.join(path2, apitype), fileapi_b1)
                                m1 = Compare_csv_diff().file_md5(apifilenm1)
                                m2 = Compare_csv_diff().file_md5(apifilenm2)
                                if m1 == m2:  # 文件md5转码值相等，不用对比操作内容
                                    same_count = same_count + 1
                                else:
                                    different_count = different_count + 1
                                    # csv文件内容不一致就比对内容输出报告
                                    # 读setting.py配置diff_html，diff_csv
                                    if setting.diff_csv == 1:  # csv文件比对
                                        # 测试报告输出到report\
                                        diff_fnm = str(apifilenm1.split('\\')[-1])
                                        diff_path_fm = rept_path + str(diff_fnm.split('.')[0] + 'diff.xlsx')
                                        # 创建diff_execl文件作为输出测试报告文件
                                        wb = openpyxl.Workbook()
                                        # ws = wb.active
                                        # sheet = wb.create_sheet('sheet', 0)  # 创建一个sheet，用于存放数据
                                        wb.save(diff_path_fm)  # 保存到excel文件
                                        version = str(diff_fnm.split('_')[0])  # 获取version
                                        key = Compare_csv_diff().get_apiversion_key(apitype, version)
                                        Compare_csv_diff().data_compy(apifilenm1, apifilenm2, diff_path_fm, key)

            # res_str=f'{apitype}接口csv文件比对一致的用例数有{same_count}，不一致的用例数有{different_count}'
            # with open(result_log_dir, 'a') as file_handle:  # .txt可以不自己新建,代码会自动新建
            #     file_handle.write(res_str)  # 写入
            #     file_handle.write('\n')

            log_json = copy(self.logTemplate)
            log_json['_dataTime'] = datetime.datetime.now().__str__()
            log_json['_metric'] = f'{apitype}接口csv文件比对一致的用例数有{same_count}，不一致的用例数有{different_count}'
            self.logUtils.logger.info(json.dumps(log_json, ensure_ascii=False))

        return out_path


    # 文件转md5码
    def file_md5(self, path):
        file_md5 = hashlib.md5()
        size = os.path.getsize(path)
        with open(path, mode='rb') as f:
            while size > 0:
                if size > 1024:
                    read_size = 1024
                else:
                    read_size = size
                file_md5.update(f.read(read_size))
                size -= read_size
        # print(file_md5.hexdigest())
        return file_md5.hexdigest()



    # 读取config/ver_key_api.yml
    def get_apiversion_key(self, interface_type, version):
            y = YamlUtils(os.path.join(BASE_DATA_DIR, 'data_test/mds_respond/ver_key_api.yml'))
            yamldict = y.get_data()
            # 将接口类型，版本号，key写入verkey_yml_path文件中
            api_list = yamldict[
                interface_type]  # 接口类型下的[{"version":"v1","key":"code"},{"version":"v2","key":"mktcode"}]
            for i in api_list:
                if i['version'] == version:
                    return i['key']



    # 根据路径和类型，返回接口类型文件夹下所有csv文件
    def return_api_csv(self, path, apitype):
            filepath = os.path.join(path, apitype)
            api_listcsv = os.listdir(filepath)
            return api_listcsv



    # 读取csv文件
    def data_compy(self,apifnm1,apifnm2,report_diff_fnm,key):
        df1 = pd.read_csv(apifnm1, encoding='utf-8', engine='c')
        df2 = pd.read_csv(apifnm2, encoding='utf-8', engine='c')
        compare = datacompy.Compare(df1, df2, join_columns=[key])
        df1_unq_rows = compare.df1_unq_rows
        df2_unq_rows = compare.df2_unq_rows
        writer = pd.ExcelWriter(report_diff_fnm, engine='xlsxwriter')
        compare.all_mismatch().to_excel(writer, sheet_name='不一致')
        df1_unq_rows.to_excel(writer, sheet_name='df1缺少的数据')
        df2_unq_rows.to_excel(writer, sheet_name='df2缺少的数据')
        writer.save()
        #writer.close()
        # openpyxl 实现xlsx的单元格颜色填充
        Compare_csv_diff().openpyxl_set_xlsx_color(report_diff_fnm)




    # openpyxl 实现xlsx的单元格颜色填充
    def openpyxl_set_xlsx_color(self, diff_file):
        red_fill = PatternFill(patternType='solid', fgColor='FF0000')  # 红色
        wb = openpyxl.load_workbook(diff_file)
        ws = wb.active
        nrows = ws.max_row
        ncols = ws.max_column
        print("nrows=",nrows)
        if nrows==2:
            for j in range(3, ncols, 2):
                if (str(ws.cell(row=2, column=j).value) != str(ws.cell(row=2, column=j + 1).value)):
                    ws.cell(row=1, column=j).fill = red_fill
                    ws.cell(row=1, column=j + 1).fill = red_fill
                    ws.cell(row=2, column=j).fill = red_fill
                    ws.cell(row=2, column=j + 1).fill = red_fill

        else:
            for i in range(2, nrows):
                for j in range(3, ncols, 2):
                    if (str(ws.cell(row=i, column=j).value) != str(ws.cell(row=i, column=j + 1).value)):
                        ws.cell(row=1, column=j).fill = red_fill
                        ws.cell(row=1, column=j + 1).fill = red_fill
                        ws.cell(row=i, column=j).fill = red_fill
                        ws.cell(row=i, column=j + 1).fill = red_fill
        wb.save(diff_file)


    def color_volume(self,volumfilediff):
        red_fill = PatternFill(patternType='solid', fgColor='FF0000')  # 红色
        yellow_fill=PatternFill(patternType='solid', fgColor='FFFF00')  # 黄色
        wb = openpyxl.load_workbook(volumfilediff)
        ws = wb.active
        nrows = ws.max_row
        for i in range(2, nrows):
            # 判断每行的snap volume 和line sum volume是否值一样，不一样高亮显示
            if str(ws.cell(row=i, column=4).value) != str(ws.cell(row=i, column=5).value):
                #print(str(ws.cell(row=2, column=4).value))
                ws.cell(row=1, column=4).fill= red_fill
                ws.cell(row=i, column=4).fill=red_fill
            if str(ws.cell(row=i, column=4).value) != str(ws.cell(row=i, column=6).value):
                ws.cell(row=1, column=6).fill=red_fill
                ws.cell(row=i, column=6).fill=red_fill
        for j in range(2, nrows, 2):
            if str(ws.cell(row=j, column=4).value) != str(ws.cell(row=j+1, column=4).value):
                ws.cell(row=1, column=4).fill=yellow_fill
                ws.cell(row=j, column=4).fill=yellow_fill
                ws.cell(row=j+1, column=4).fill=yellow_fill
        wb.save(volumfilediff)


if __name__=='__main__':
    #多进程
    #Compare_csv_diff().sh1_compare_comm_fun()
    #单线程
    Compare_csv_diff().act_diff_case()






